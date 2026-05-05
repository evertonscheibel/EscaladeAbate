import json
import sys
import os
from openpyxl import load_workbook

def export_sif(template_path, output_path, data_json):
    try:
        # Carregar dados
        import json
        data = json.loads(data_json)
        
        # Carregar template com macros preservadas
        wb = load_workbook(template_path, keep_vba=True)
        
        # Supondo que a aba principal se chama 'Planilha1' ou similar
        # Ajustar conforme o template real
        ws = wb.active
        
        # Preencher Cabeçalho (Exemplo de células hipotéticas)
        # ws['B5'] = data.get('header', {}).get('slaughterDate', '')
        # ws['B6'] = data.get('header', {}).get('veterinarian', '')
        
        # Preencher Linhas
        start_row = 10 # Linha onde começam os lotes
        for i, line in enumerate(data.get('lines', [])):
            row = start_row + i
            ws.cell(row=row, column=1, value=line.get('sequence'))
            ws.cell(row=row, column=2, value=line.get('producerName'))
            ws.cell(row=row, column=3, value=line.get('municipio'))
            ws.cell(row=row, column=4, value=line.get('curral'))
            ws.cell(row=row, column=5, value=line.get('boi', 0))
            ws.cell(row=row, column=6, value=line.get('vaca', 0))
            ws.cell(row=row, column=7, value=line.get('novilha', 0))
            ws.cell(row=row, column=8, value=line.get('bubalino', 0))
            ws.cell(row=row, column=9, value=line.get('touro', 0))
            ws.cell(row=row, column=10, value=line.get('total', 0))
            ws.cell(row=row, column=11, value=line.get('nf', ''))
            ws.cell(row=row, column=12, value=line.get('gta', ''))
            
        # Salvar resultado
        wb.save(output_path)
        print(f"SUCCESS: {output_path}")
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python export_sif_xlsm.py template_path output_path data_json")
        sys.exit(1)
        
    export_sif(sys.argv[1], sys.argv[2], sys.argv[3])
